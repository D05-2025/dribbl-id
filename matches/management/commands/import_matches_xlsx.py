from pathlib import Path
from datetime import datetime
from django.core.management.base import BaseCommand, CommandError
from django.utils.timezone import make_aware, is_naive
from openpyxl import load_workbook
from matches.models import Match

class Command(BaseCommand):
    help = "Import dataset matches from Excel (.xlsx)"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to Excel file")
        parser.add_argument("--sheet", type=str, default=None, help="Nama sheet (opsional)")
        parser.add_argument(
            "--dtfmt",
            type=str,
            default="%Y-%m-%d %H:%M:%S",
            help="Format waktu jika bukan datetime Excel",
        )

    def handle(self, *args, **opt):
        path = Path(opt["xlsx_path"])
        if not path.exists():
            raise CommandError(f"File tidak ditemukan: {path}")

        wb = load_workbook(filename=path, data_only=True)
        ws = wb[opt["sheet"]] if opt["sheet"] else wb.active

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: headers.index(name) for name in headers}

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue

            home_team = str(row[idx["home_team"]]).strip()
            away_team = str(row[idx["away_team"]]).strip()
            tipoff_raw = row[idx["tipoff_at"]]

            # Convert tipoff_at
            if isinstance(tipoff_raw, datetime):
                tipoff_at = tipoff_raw
            else:
                tipoff_at = datetime.strptime(str(tipoff_raw).strip(), opt["dtfmt"])
            if is_naive(tipoff_at):
                tipoff_at = make_aware(tipoff_at)

            defaults = {
                "venue": row[idx.get("venue")] or None,
                "status": str(row[idx.get("status")] or "scheduled"),
                "home_score": int(row[idx.get("home_score")] or 0),
                "away_score": int(row[idx.get("away_score")] or 0),
                "image_url": row[idx.get("image_url")] or None,
            }

            Match.objects.update_or_create(
                home_team=home_team,
                away_team=away_team,
                tipoff_at=tipoff_at,
                defaults=defaults,
            )
            count += 1

        self.stdout.write(self.style.SUCCESS(f"{count} matches berhasil diimport."))
